# https://www.youtube.com/watch?v=bu5wXjz2KvU

import pandas as pd
import numpy as np
import os
import datetime
import gspread
import streamlit as st
import time
import zipfile

from datetime import datetime
from datetime import timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
from PIL import Image

import psycopg2  # pip install psycopg2
import psycopg2.extras 
from psycopg2.extras import execute_values
from google.oauth2 import service_account

DB_HOST = "database-1.cdcogkfzajf0.us-east-1.rds.amazonaws.com"
DB_NAME = "postgres"
DB_USER = "postgres"
DB_PASS = "15512332"

###### CONECTANDO PLANILHAS ##########

# Connect to Google Sheets
service_account_info = st.secrets["GOOGLE_SERVICE_ACCOUNT"]

scope = ['https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive"]

credentials = service_account.Credentials.from_service_account_info(service_account_info, scopes=scope)

sa = gspread.authorize(credentials)

st.title('Gerador de Ordem de Produção')

st.write("Base para gerar as ordens de produção")
st.write("https://docs.google.com/spreadsheets/d/18ZXL8n47qSLFLVO5tBj7-ADpqmMyFwCgs4cxxtBB9Xo/edit#gid=0")

st.write("Planilha que guarda ordens geradas")
st.write("https://docs.google.com/spreadsheets/d/1IOgFhVTBtlHNBG899QqwlqxYlMcucTx74zRA29YBHKA/edit#gid=1228486917")

st.write("Planilha para impressão de etiquetas")
st.write("https://docs.google.com/spreadsheets/d/1jojKHPBKeALheutJyphsPS-LGNu1e2BC54AAqRnF-us/edit?gid=1389272651#gid=1389272651")

name_sheet = 'Bases para sequenciamento'

worksheet1 = 'Base_Carretas'
worksheet2 = 'Carga_Vendas'

# filename = r"C:\Users\pcp2\ordem de producao\Ordem-de-producao\service_account.json"
# filename = "service_account.json"

# sa = gspread.service_account(filename)
sh = sa.open(name_sheet)

@st.cache_data(ttl=600)
def get_data_from_sheets():
    """Carrega os dados das planilhas e retorna como DataFrames."""
    wks1 = sh.worksheet(worksheet1)
    wks2 = sh.worksheet(worksheet2)

    list1 = wks1.get_all_records()
    list2 = wks2.get_all_records()

    # Transformando em dataframes
    base_carretas = pd.DataFrame(list1)
    base_carga = pd.DataFrame(list2)

    return base_carretas, base_carga

###### TRATANDO DADOS #########

##### Tratando datas######

base_carretas, base_carga = get_data_from_sheets()

dados_carga = base_carga
dados_carreta = base_carretas

dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].apply(lambda x: "0" + str(x) if len(str(x))==5 else str(x))
dados_carreta['Recurso'] = dados_carreta['Recurso'].apply(lambda x: "0" + str(x) if len(str(x))==5 else str(x))

base_carga = base_carga[['PED_PREVISAOEMISSAODOC','PED_RECURSO.CODIGO', 'PED_QUANTIDADE']]
base_carga['PED_PREVISAOEMISSAODOC'] = pd.to_datetime(
    base_carga['PED_PREVISAOEMISSAODOC'], format='%d/%m/%Y', errors='coerce')
base_carga['Ano'] = base_carga['PED_PREVISAOEMISSAODOC'].dt.strftime('%Y')
base_carga['PED_PREVISAOEMISSAODOC'] = base_carga.PED_PREVISAOEMISSAODOC.dt.strftime(
    '%d/%m/%Y')

#### renomeando colunas#####

base_carga = base_carga.rename(columns={'PED_PREVISAOEMISSAODOC': 'Datas',
                                        'PED_RECURSO.CODIGO': 'Recurso',
                                        'PED_QUANTIDADE': 'Qtde'})

##### Valores nulos######

base_carga.dropna(inplace=True)
base_carga.reset_index(drop=True)

# base_carga[base_carga['Recurso'] == '034830CO']

today = datetime.now()
ts = pd.Timestamp(today)
today = today.strftime('%d/%m/%Y')

filenames = []

def consultar_carretas(data_inicial, data_final, dados_carga, dados_carreta):

    sufixos_para_remover = ['AV', 'VM', 'VJ', 'AN', 'AS', 'CO', 'LC', ]

    dados_carga['PED_RECURSO.CODIGO'] = dados_carga['PED_RECURSO.CODIGO'].apply(
        lambda x: x[:-2].rstrip() if str(x)[-2:] in sufixos_para_remover else x
    )

    dados_carga['PED_PREVISAOEMISSAODOC'] = pd.to_datetime(
        dados_carga['PED_PREVISAOEMISSAODOC'], 
        format="%d/%m/%Y",  # Dia/Mês/Ano
        errors='coerce'  # Caso queira tratar erros, como datas inválidas
    )

    dados_carga_data_filtrada = dados_carga[(dados_carga['PED_PREVISAOEMISSAODOC'] >= data_inicial) & (dados_carga['PED_PREVISAOEMISSAODOC'] <= data_final)]

    dados_carga_data_filtrada['PED_QUANTIDADE'] = dados_carga_data_filtrada['PED_QUANTIDADE'].astype(float)

    carretas_unica = dados_carga_data_filtrada[['PED_PREVISAOEMISSAODOC','PED_RECURSO.CODIGO','PED_QUANTIDADE']]
    agrupado = carretas_unica.groupby(['PED_PREVISAOEMISSAODOC','PED_RECURSO.CODIGO'])['PED_QUANTIDADE'].sum().reset_index()

    dados_carreta['Recurso'] = dados_carreta['Recurso'].apply(lambda x: "0" + str(x) if len(str(x))==5 else x)

    agrupado['Contém'] = agrupado['PED_RECURSO.CODIGO'].apply(
        lambda x: '✅' if x in dados_carreta['Recurso'].astype(str).values else '❌'
    )

    resultado = agrupado.values.tolist()

    return resultado

def criar_array_datas(data_inicial, data_final):
    # Converte as strings de data para objetos datetime
    data_inicial = data_inicial.strftime("%d/%m/%Y")
    data_final = data_final.strftime("%d/%m/%Y")
    data_inicial = datetime.strptime(data_inicial, "%d/%m/%Y")
    data_final = datetime.strptime(data_final, "%d/%m/%Y")
    
    # Gera a lista de datas
    array_datas = []
    data_atual = data_inicial
    
    while data_atual <= data_final:
        # Adiciona a data atual formatada como string à lista
        array_datas.append(data_atual.strftime("%d/%m/%Y"))
        # Avança um dia
        data_atual += timedelta(days=1)
    
    return array_datas

def gerar_etiquetas_montagem(tipo_filtro,df):
    
    # tab_completa_montagem[tab_completa_montagem['Célula'] == 'CHASSI']
    # tab_completa[tab_completa['Célula'] == 'CHASSI']
    df=df[df['Célula'] == 'CHASSI']
    df['cor'] = 'Cinza'

    # Criar uma coluna adicional
    # Repetir as linhas de acordo com a quantidade total
    df = df.loc[df.index.repeat(df['Qtde_total'])].reset_index(drop=True)
    df['sequencia'] = ''
    df = df.sort_values(by=['Célula','Código']).reset_index(drop=True)
    # df['Célula'].iloc[:,4:]
    contador = 1
    
    for i in range(len(df)):
        
        try:
            if df['Código'][i] == df['Código'][i-1]:
                df['sequencia'][i] = str(contador) + "/" + str(df['Qtde_total'][i])
                contador += 1
            else:
                contador = 1
                df['sequencia'][i] = str(contador) + "/" + str(df['Qtde_total'][i])
                contador += 1
        except:
            df['sequencia'][i] = str(contador) + "/" + str(df['Qtde_total'][i])
            contador += 1
            continue

    codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[8:10]
    
    # df2 = df.groupby('Célula', as_index=False).apply(lambda x: x.append(pd.Series(name=x.name))).reset_index(drop=True)

    df['codificacao'] = df.apply(criar_codificacao, axis=1, codigo_unico=codigo_unico)

    df['Concatenacao'] = df.apply(lambda row: f"{row['Código']} - {row['Peca']} ☐\n{row['Código_y']} - {row['Peca_y']} ☐       {row['codificacao']}\nCélula: {row['Célula']} Quantidade: {row['sequencia']}        \nCor: Cinza\nMontagem:__________Data:__________\nSolda:__________Data:__________\nPintura:__________Data:__________", axis=1)
    
    # for i in range(len(df)):
    #     try:

    # Crie um novo DataFrame com as linhas em branco e o valor da data na última linha
    # new_rows = []
    # for index, row in df.iterrows():
    #     new_rows.append(row)
    #     if index < len(df) - 1 and df.at[index, 'Célula'] != df.at[index + 1, 'Célula']:
    #         new_rows.append(pd.Series(['']*13, index=df.columns))
    #     elif index == len(df) - 1:
    #         new_rows.append(pd.Series(['']*13, index=df.columns))
    
    # df = pd.DataFrame(new_rows).reset_index(drop=True)

    # # Adicionar linha em branco ao final de cada grupo

    # Seus valores a serem anexados
    # valores = df['Concatenacao'].tolist()

    # # Separar valores pares e ímpares
    # valores_pares = valores[::2]
    # valores_impares = valores[1::2]

    # # Anexar à planilha
    # intervalo_pares = "etiquetas!A3:A" + str(len(valores_pares)+3)  # +3 para ajustar a contagem de linhas
    # intervalo_impares = "etiquetas!B3:B" + str(len(valores_impares)+3)  # +3 para ajustar a contagem de linhas

    # planilha.values_clear("etiquetas!A:B")

    # # Anexar valores pares
    # planilha.values_append(intervalo_pares, {'valueInputOption': 'RAW'}, {'values': [[valor] for valor in valores_pares]})

    # # Anexar valores ímpares
    # planilha.values_append(intervalo_impares, {'valueInputOption': 'RAW'}, {'values': [[valor] for valor in valores_impares]})
    
    # # Salvar as alterações no Excel
    # wb.save('etiquetas.xlsx')
    # wb.close()

    # my_file = "etiquetas.xlsx"

    # return my_file
    return df

def gerar_etiquetas(tipo_filtro,df,df_montagem):

    
    # tab_completa_montagem[tab_completa_montagem['Célula'] == 'CHASSI']
    # tab_completa[tab_completa['Célula'] == 'CHASSI']

    # Abra a planilha pelo seu URL (key)
    planilha = sa.open_by_key("1jojKHPBKeALheutJyphsPS-LGNu1e2BC54AAqRnF-us")

    # Acesse a aba desejada
    aba = planilha.worksheet("etiquetas")

    # Criar uma coluna adicional
    # Repetir as linhas de acordo com a quantidade total
    df = df.loc[df.index.repeat(df['Qtde_total'])].reset_index(drop=True)
    df['sequencia'] = ''
    df = df.sort_values(by=['Célula','Recurso_cor']).reset_index(drop=True)
    # df['Célula'].iloc[:,4:]
    contador = 1
    
    for i in range(len(df)):
        
        try:
            if df['Código'][i] == df['Código'][i-1]:
                df['sequencia'][i] = str(contador) + "/" + str(df['Qtde_total'][i])
                contador += 1
            else:
                contador = 1
                df['sequencia'][i] = str(contador) + "/" + str(df['Qtde_total'][i])
                contador += 1
        except:
            df['sequencia'][i] = str(contador) + "/" + str(df['Qtde_total'][i])
            contador += 1
            continue

    codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[8:10]
    
    # df2 = df.groupby('Célula', as_index=False).apply(lambda x: x.append(pd.Series(name=x.name))).reset_index(drop=True)

    df['codificacao'] = df.apply(criar_codificacao, axis=1, codigo_unico=codigo_unico)
    
    # df['Concatenacao'] = df.apply(lambda row: f"{row['Código']} - {row['Peca']}     {row['codificacao']}\nCélula: {row['Célula']} Quantidade: {row['sequencia']}        \nCor: ☐Azul  ☐Amarelo  ☐Cinza  ☐Laranja  ☐Verde  ☐Vermelho\nMontagem:__________Data:__________\nSolda:__________Data:__________\nPintura:__________Data:__________" if row['cor'] != 'Cinza' else f"{row['Código']} - {row['Peca']}     {row['codificacao']}\nCélula: {row['Célula']} Quantidade: {row['sequencia']}        \nCor: {row['cor']}\nMontagem:__________Data:__________\nSolda:__________Data:__________\nPintura:__________Data:__________", axis=1)
    df['Concatenacao'] = df.apply(lambda row: f"{row['Código']} - {row['Peca'][:30]}\n{row['codificacao']}\nCélula: {row['Célula']} Quantidade: {row['sequencia']}        \nCor: {row['cor']}\nData:{tipo_filtro}", axis=1)

    # df_etiquetas_montagem = gerar_etiquetas_montagem(tipo_filtro,df_montagem)

    # df_final = pd.concat([df,df_etiquetas_montagem]).reset_index(drop=True)
    
    # Crie um novo DataFrame com as linhas em branco e o valor da data na última linha
    # new_rows = []
    # for index, row in df_final.iterrows():
    #     new_rows.append(row)
    #     if index < len(df_final) - 1 and df_final.at[index, 'Célula'] != df_final.at[index + 1, 'Célula']:
    #         new_rows.append(pd.Series(['']*12, index=df_final.columns))
    #         new_rows.append(pd.Series(['']*12, index=df_final.columns))
    #     elif index == len(df_final) - 1:
    #         new_rows.append(pd.Series(['']*12, index=df_final.columns))
    #         new_rows.append(pd.Series(['']*12, index=df_final.columns))
    
    # df_final = pd.DataFrame(new_rows).reset_index(drop=True)

    # Adicionar linha em branco ao final de cada grupo

    # Seus valores a serem anexados
    valores = df['Concatenacao'].tolist()

    # Separar valores pares e ímpares
    valores_pares = valores[::2]
    valores_impares = valores[1::2]

    # Anexar à planilha
    intervalo_pares = "etiquetas!A3:A" + str(len(valores_pares)+3)  # +3 para ajustar a contagem de linhas
    intervalo_impares = "etiquetas!B3:B" + str(len(valores_impares)+3)  # +3 para ajustar a contagem de linhas

    planilha.values_clear("etiquetas!A:B")

    # Anexar valores pares
    planilha.values_append(intervalo_pares, {'valueInputOption': 'RAW'}, {'values': [[valor] for valor in valores_pares]})

    # Anexar valores ímpares
    planilha.values_append(intervalo_impares, {'valueInputOption': 'RAW'}, {'values': [[valor] for valor in valores_impares]})
    
def criar_codificacao(row, codigo_unico):

    if row['Célula'] == "EIXO COMPLETO":
        return row['Célula'][0:3] + codigo_unico + "C"
    elif row['Célula'] == "EIXO SIMPLES":
        return row['Célula'][0:3] + codigo_unico + "S"
    else:
        return row['Célula'][0:3] + codigo_unico

def str_to_float(stringNumber):
    """Função para transformar string em float"""

    transformed = stringNumber.replace(",",".")

    return float(transformed)
# Lendo tabela com consumo de cores

# Abra a planilha pelo seu URL (key)
planilha = sa.open_by_key("1RJH3k5brgO3nmEQPNOKdp_HFqVbJMUcwGEQ1XQyHtCs")

# Acesse a aba desejada
aba = planilha.worksheet("CONSUMO PU")
valores = aba.get()

lista_columns_consumo = valores[0]
valores = valores[1:]

df_consumo_pu = pd.DataFrame(columns=lista_columns_consumo, data=valores)
df_consumo_pu['Consumo Pó (kg)'] = df_consumo_pu['Consumo Pó (kg)'].apply(str_to_float)
df_consumo_pu["Consumo PU (L)"] = df_consumo_pu["Consumo PU (L)"].apply(str_to_float)
df_consumo_pu["Consumo Catalisador (L)"] = df_consumo_pu["Consumo Catalisador (L)"].apply(str_to_float)

def unique(list1):
    x = np.array(list1)
    print(np.unique(x))

with st.sidebar:

    image = Image.open('logo-cemagL.png')
    st.image(image, width=300)

# tipo_filtro = st.date_input('Data: ')
data = datetime.strptime('2025-03-26', '%Y-%m-%d').date()

tipo_filtro = st.date_input(
    'Data:',
    # value=(datetime.now(), datetime.now())
    value=(data, data)
)

data_inicio, data_fim = tipo_filtro
resultado = criar_array_datas(data_inicio, data_fim)
data_inicio = data_inicio.strftime("%Y-%m-%d")
data_fim_consulta = data_fim.strftime("%Y-%m-%d")
data_fim = data_fim.strftime("%d/%m/%Y")
carretas_na_base = consultar_carretas(data_inicio,data_fim_consulta, dados_carga, dados_carreta)
df = pd.DataFrame(carretas_na_base, columns=['Data','Código Carreta','Quantidade','Contém'])
df['Data'] = pd.to_datetime(df['Data']).dt.date

# Exibir a tabela no Streamlit
st.write("Tabela de Carretas:")
st.dataframe(df[['Data','Código Carreta','Quantidade','Contém']])

values = ['Selecione','Pintura','Montagem','Solda', 'Serralheria', 'Carpintaria', 'Etiquetas']
setor = st.selectbox('Escolha o setor', values)

check_atualizar_base_carga = st.checkbox('Atualizar base com novos dados?', value=False)

submit_button = st.button(label='Gerar')

def insert_pintura(data_carga, dados, check_atualizar_base_carga):
    
    # data_carga = datetime.strptime(data_carga,'%d/%m/%Y').strftime('%Y-%m-%d')

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Verifica se existe registros com a data atual
    sql_select = 'SELECT * FROM pcp.gerador_ordens_pintura where data_carga = %s'
    cur.execute(sql_select,(data_carga,))
    dados_existente = cur.fetchall()

    if len(dados_existente) > 0:
        if check_atualizar_base_carga:
            # Exclui os registros com a data_carga fornecida
            sql_delete = 'DELETE FROM pcp.gerador_ordens_pintura WHERE data_carga = %s::date;'
            cur.execute(sql_delete, (data_carga,))
            
            conn.commit()

            for dado in dados:
                # Construir e executar a consulta INSERT
                query = ("INSERT INTO pcp.gerador_ordens_pintura (celula, codigo, peca, cor, qt_planejada, data_carga) VALUES (%s, %s, %s, %s, %s, %s::date)")
                cur.execute(query, dado)

            # Commit para aplicar as alterações
            conn.commit()
    else:

        # Exclui os registros com a data_carga fornecida
        sql_delete = 'DELETE FROM pcp.gerador_ordens_pintura WHERE data_carga = %s::date;'
        cur.execute(sql_delete, (data_carga,))
        
        print(sql_delete)

        conn.commit()

        for dado in dados:
            # Construir e executar a consulta INSERT
            query = ("INSERT INTO pcp.gerador_ordens_pintura (celula, codigo, peca, cor, qt_planejada, data_carga) VALUES (%s, %s, %s, %s, %s, %s::date)")
            cur.execute(query, dado)

        # Commit para aplicar as alterações
        conn.commit()

def insert_montagem(data_carga, dados, check_atualizar_base_carga):
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Verifica se existe registros com a data atual
    sql_select = 'SELECT * FROM pcp.gerador_ordens_montagem where data_carga = %s'
    cur.execute(sql_select,(data_carga,))
    dados_existente = cur.fetchall()

    if len(dados_existente) > 0:
        if check_atualizar_base_carga:

            # Exclui os registros com a data_carga fornecida
            sql_delete = 'DELETE FROM pcp.gerador_ordens_montagem WHERE data_carga = %s;'
            cur.execute(sql_delete, (data_carga,))
            conn.commit()

            for dado in dados:
                # Construir e executar a consulta INSERT
                query = ("INSERT INTO pcp.gerador_ordens_montagem (celula, codigo, peca, qt_planejada, data_carga) VALUES (%s, %s, %s, %s, %s)")
                cur.execute(query, dado)

            # Commit para aplicar as alterações
            conn.commit()

    else:

        # Exclui os registros com a data_carga fornecida
        sql_delete = 'DELETE FROM pcp.gerador_ordens_montagem WHERE data_carga = %s;'
        cur.execute(sql_delete, (data_carga,))
        conn.commit()

        for dado in dados:
            # Construir e executar a consulta INSERT
            query = ("INSERT INTO pcp.gerador_ordens_montagem (celula, codigo, peca, qt_planejada, data_carga) VALUES (%s, %s, %s, %s, %s)")
            cur.execute(query, dado)

        # Commit para aplicar as alterações
        conn.commit()

def tratar_conjuntos_iguais(base_carretas,base_carga):

    base_carreta_montagem = base_carretas.copy()
    base_carga_montagem = base_carga.copy()

    ########################### Montagem ###########################

    base_carreta_montagem['Código'] = base_carreta_montagem['Código'].astype(str)
    base_carreta_montagem['Recurso'] = base_carreta_montagem['Recurso'].astype(str)

    ####### retirando cores dos códigos######

    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].astype(str)

    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('AM', '')
    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('AN', '')
    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('VJ', '')
    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('LC', '')
    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('VM', '')
    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('AV', '')

    # base_carga_montagem[base_carga_montagem['Datas'] == '01/06/2023']

    ###### retirando espaco em branco####

    base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.strip()

    base_carreta_montagem.drop(['Etapa3', 'Etapa4',
                           'Etapa5'], axis=1, inplace=True)

    base_carreta_montagem_chassi = base_carreta_montagem[base_carreta_montagem['Célula'] == 'CHASSI']
    # base_carreta_montagem[base_carreta_montagem['Célula'] == 'EIXO SIMPLES']

    chassi_pintura = base_carreta_montagem_chassi[base_carreta_montagem_chassi['Etapa2'] != ''].drop(columns={'Etapa'})
    chassi_montagem = base_carreta_montagem_chassi[base_carreta_montagem_chassi['Etapa'] != ''].drop(columns={'Etapa2'})

    chassis_join = chassi_montagem.merge(chassi_pintura, how='left', on='Recurso')

    ##################################################################################

    escolha_data = (base_carga_montagem['Datas'] == tipo_filtro)
    filtro_data = base_carga_montagem.loc[escolha_data]
    filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

    filtro_data = filtro_data.reset_index(drop=True)
    filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

    for i in range(len(filtro_data)):
        if filtro_data['Recurso'][i][0] == '0':
            filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]

    ##### juntando planilhas de acordo com o recurso#######

    tab_completa = pd.merge(filtro_data, chassis_join[[
                            'Recurso', 'Código_x', 'Peca_x', 'Qtde_x', 'Célula_x', 'Código_y', 'Peca_y']], on=['Recurso'], how='left')
    # tab_completa = tab_completa.dropna(axis=0)

    tab_completa = tab_completa.groupby(['Datas','Código_x','Peca_x','Código_y', 'Peca_y','Célula_x']).sum(['Qtde','Qtde_x']).reset_index()

    for i in range(len(tab_completa)):

        if len(tab_completa['Código_x'][i]) == 5:
            tab_completa['Código_x'][i] = '0' + tab_completa['Código_x'][i]   

        if len(tab_completa['Código_y'][i]) == 5:
            tab_completa['Código_y'][i] = '0' + tab_completa['Código_y'][i]   

    chassi = tab_completa.rename(columns={'Código_x':'Código','Peca_x':'Peca','Qtde':'Qtde_total','Célula_x':'Célula'})

    return chassi

if submit_button:
    base_carretas_original = base_carretas.copy()
    base_carga_original = base_carga.copy()
    for idx, tipo_filtro in enumerate(resultado):
        data_nome_planilha = tipo_filtro.replace("/","-")[:5]
        base_carretas = base_carretas_original.copy()
        base_carga = base_carga_original.copy()

        if setor == 'Pintura':

            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            base_carretas.drop(['Etapa', 'Etapa3', 'Etapa4'], axis=1, inplace=True)

            base_carretas.drop(
                base_carretas[(base_carretas['Etapa2'] == '')].index, inplace=True)

            base_carretas = base_carretas.reset_index(drop=True)

            base_carretas = base_carretas.astype(str)

            for d in range(0, base_carretas.shape[0]):

                if len(base_carretas['Código'][d]) == 5:
                    base_carretas['Código'][d] = '0' + base_carretas['Código'][d]

            # separando string por "-" e adicionando no dataframe antigo

            base_carga["Recurso"] = base_carga["Recurso"].astype(str)

            tratando_coluna = base_carga["Recurso"].str.split(
                " - ", n=1, expand=True)

            base_carga['Recurso'] = tratando_coluna[0]

            # tratando cores da string

            base_carga['Recurso_cor'] = base_carga['Recurso']

            base_carga = base_carga.reset_index(drop=True)

            df_cores = pd.DataFrame({'Recurso_cor': ['AN', 'VJ', 'LJ', 'LC', 'VM', 'AV', 'sem_cor', 'AS', 'CO'],
                                    'cor': ['Azul', 'Verde', 'Laranja Jacto', 'Laranja', 'Vermelho', 'Amarelo', 'Laranja', 'Azul Sm.', 'Cinza']})

            nome_cor_para_sigla = dict(zip(df_cores['cor'], df_cores['Recurso_cor']))

            cores = ['AM', 'AN', 'VJ', 'LJ', 'LC', 'VM', 'AV', 'AS', 'CO']

            base_carga = base_carga.astype(str)

            for r in range(0, base_carga.shape[0]):
                base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][len(
                    base_carga['Recurso_cor'][r])-3:len(base_carga['Recurso_cor'][r])]
                base_carga['Recurso_cor'] = base_carga['Recurso_cor'].str.strip()

                if len(base_carga['Recurso_cor'][r]) > 2:
                    base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][1:3]

                if base_carga['Recurso_cor'][r] not in cores:
                    base_carga['Recurso_cor'][r] = "LC"

            base_carga = pd.merge(base_carga, df_cores, on=[
                                'Recurso_cor'], how='left')

            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'AN', '')  # Azul
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'VJ', '')  # Verde
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'LC', '')  # Laranja
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'VM', '')  # Vermelho
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'AV', '')  # Amarelo
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'AS', '')  # Amarelo

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == tipo_filtro)
            filtro_data = base_carga.loc[escolha_data]
            # filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            # procv e trazendo as colunas que quero ver

            filtro_data = filtro_data.reset_index(drop=True)

            for i in range(len(filtro_data)):
                if filtro_data['Recurso'][i][0] == '0':
                    filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]

            # tab_completa['Recurso'] = tab_completa['Recurso'].apply(lambda x: "0" + str(x) if len(str(x)) == 5 else x)
            filtro_data['Recurso'] = filtro_data['Recurso'].apply(lambda x: "0" + str(x)  if len(str(x)) == 5 else x)

            tab_completa = pd.merge(filtro_data, base_carretas, on=[
                                    'Recurso'], how='left')

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa = tab_completa.reset_index(drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(drop=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # tratando coluna de código

            for t in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][t]) == 5:
                    tab_completa['Código'][t] = '0' + \
                        tab_completa['Código'][t][0:5]

                if len(tab_completa['Código'][t]) == 8:
                    tab_completa['Código'][t] = tab_completa['Código'][t][0:6]

            # criando coluna de quantidade total de itens

            tab_completa = tab_completa.dropna()

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(',', '.')

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa = tab_completa.dropna(axis=0)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']
            print(tab_completa)
            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y', 'LEAD TIME', 'flag peça', 'Etapa2'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas', 'Recurso_cor', 'cor']).sum()
            tab_completa.reset_index(inplace=True)

            # linha abaixo exclui eixo simples do sequenciamento da pintura
            # tab_completa.drop(tab_completa.loc[tab_completa['Célula']=='EIXO SIMPLES'].index, inplace=True)
            tab_completa.reset_index(inplace=True, drop=True)

            tab_completa['Etapa5'].unique()

            # for t in range(0, len(tab_completa)):

                # if tab_completa['Célula'][t] == 'FUEIRO' or \
                #         tab_completa['Célula'][t] == 'LATERAL' or \
                #         tab_completa['Célula'][t] == 'PLAT. TANQUE. CAÇAM.':

                #     tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + \
                #         tab_completa['Recurso_cor'][t]

                # else:

                #     tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + 'CO'
                #     tab_completa['cor'][t] = 'Cinza'

            # contem_cinza = tab_completa['Etapa5'].str.contains('CINZA')
            # tab_completa.loc[contem_cinza, 'Etapa5'] = 'CINZA'

            # contem_colorido = tab_completa['Etapa5'].str.contains('COLORIDO')
            # tab_completa.loc[contem_colorido, 'Etapa5'] = 'COLORIDO'

            # contem_preto = tab_completa['Etapa5'].str.contains('PRETO')
            # tab_completa.loc[contem_preto, 'Etapa5'] = 'PRETO'

            # Normaliza os valores da coluna 'Etapa5' para identificar corretamente as cores
            tab_completa.loc[tab_completa['Etapa5'].str.contains('CINZA', na=False), 'Etapa5'] = 'CINZA'
            tab_completa.loc[tab_completa['Etapa5'].str.contains('COLORIDO', na=False), 'Etapa5'] = 'COLORIDO'
            tab_completa.loc[tab_completa['Etapa5'].str.contains('PRETO', na=False), 'Etapa5'] = 'PRETO'

            # Função para definir a coluna 'Recurso_cor'
            def definir_recurso_cor(row, corPlanilha, siglaCor):
                if row['Etapa5'] == corPlanilha:
                    return row['Código'] + siglaCor
                else:
                    return row['Código'] + row['Recurso_cor']

            # Função para definir a coluna 'cor'
            def definir_cor(row, corPlanilha, returnCor):
                if row['Etapa5'] == corPlanilha:
                    return returnCor
                else:
                    return row['cor']

            # Aplicando as funções corretamente
            tab_completa['Recurso_cor'] = tab_completa.apply(lambda row: definir_recurso_cor(row, 'CINZA', 'Cinza'), axis=1)
            tab_completa['cor'] = tab_completa.apply(lambda row: definir_cor(row, 'CINZA', 'Cinza'), axis=1)

            tab_completa['Recurso_cor'] = tab_completa.apply(lambda row: definir_recurso_cor(row, 'PRETO', 'Preto'), axis=1)
            tab_completa['cor'] = tab_completa.apply(lambda row: definir_cor(row, 'PRETO', 'Preto'), axis=1)

            # Consumo de tinta

            tab_completa = tab_completa.merge(df_consumo_pu[['Codigo item','Consumo Pó (kg)','Consumo PU (L)','Consumo Catalisador (L)']], left_on='Código', right_on='Codigo item', how='left').fillna(0)
            
            tab_completa['Consumo Pó (kg)'] = tab_completa['Consumo Pó (kg)'] * tab_completa['Qtde_total']
            tab_completa['Consumo PU (L)'] = tab_completa['Consumo PU (L)'] * tab_completa['Qtde_total']
            tab_completa['Consumo Catalisador (L)'] = tab_completa['Consumo Catalisador (L)'] * tab_completa['Qtde_total']

            consumo_po = sum(tab_completa['Consumo Pó (kg)'])
            consumo_po = f'{round(consumo_po / 25, 2)} caixa(s)'

            consumo_pu_litros = sum(tab_completa['Consumo Pó (kg)'])
            consumo_pu_latas = round(consumo_pu_litros / 3.08, 2)
            consumo_pu = f'{consumo_pu_latas} lata(s)'

            consumo_catalisador_litros = sum(tab_completa['Consumo Catalisador (L)'])
            consumo_catalisador_latas = round(consumo_catalisador_litros * 1000 / 400, 2)
            consumo_cata = f'{consumo_catalisador_latas} lata(s)'

            diluente = f'{round((consumo_pu_litros * 0.80) / 5, 2)} lata(s)'

            ###########################################################################################

            cor_unique = tab_completa['cor'].unique()
            if idx == 0:
                st.write("Arquivos para download")

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            tab_completa = tab_completa.reset_index(drop=True)

            # carga_unique = tab_completa['Carga'].unique()
            file_counter = 1
            rows_per_file = 21

            # for carga in carga_unique:
            for i in range(len(cor_unique)):

                start_index = 0

                filtro_excel = (tab_completa['cor'] == cor_unique[i])
                filtrar = tab_completa.loc[filtro_excel]
                filtrar = filtrar.reset_index(drop=True)
                filtrar = filtrar.groupby(
                    ['Código', 'Peca', 'Célula', 'Datas', 'Recurso_cor', 'cor']
                ).sum().reset_index()
                filtrar.sort_values(by=['Célula'], inplace=True)
                filtrar = filtrar.reset_index(drop=True)
                print(filtrar)
                while start_index < len(filtrar):
                    # Criar um novo Workbook para cada conjunto de 21 linhas
                    wb = Workbook()
                    wb = load_workbook('modelo_op_pintura.xlsx')
                    ws = wb.active

                    k = 9  # Início da linha no Excel

                    # Define o limite superior para as linhas deste arquivo
                    end_index = min(start_index + rows_per_file, len(filtrar))

                    # Escreve os dados no Excel para as linhas entre start_index e end_index
                    for j in range(start_index, end_index):
                        
                        cor_nome = filtrar['cor'][j]
                        sigla_cor = nome_cor_para_sigla.get(cor_nome, 'sem_cor')

                        ws['F5'] = cor_unique[i]  # nome da coluna é '0'
                        ws['AD5'] = datetime.now()  # data de hoje
                        ws['M4'] = tipo_filtro  # data da carga
                        ws['B' + str(k)] = f"{str(filtrar['Código'][j])}{sigla_cor}"
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        ws['K3'] = consumo_cata
                        ws['Q3'] = consumo_po
                        ws['AE3'] = consumo_pu
                        ws['AN3'] = diluente
                        k += 1

                    # Salvar o arquivo com numeração sequencial
                    file_name = f"Pintura {cor_unique[i]} {data_nome_planilha} {file_counter}.xlsx"
                    wb.save(file_name)
                    filenames.append(file_name)

                    # Incrementar índice e contador de arquivos
                    start_index = end_index
                    file_counter += 1

                
            data_insert_sql = tab_completa[['Célula','Código','Peca','cor','Qtde_total','Datas']]
            data_insert_sql = data_insert_sql.groupby(['Célula','Código','Peca','cor','Datas']).sum().reset_index()[['Célula','Código','Peca','cor','Qtde_total','Datas']]
            data_insert_sql['Datas'] = pd.to_datetime(data_insert_sql['Datas'], format='%d/%m/%Y')

            data_insert_sql = data_insert_sql.values.tolist()

            # data_formatada = datetime.strptime(datetime.strptime(tipo_filtro,'%d/%m/%Y').strftime('%Y-%m-%d'),'%Y-%m-%d').date()
            tipo_filtro = pd.to_datetime(tipo_filtro, format='%d/%m/%Y')
            insert_pintura(tipo_filtro, data_insert_sql, check_atualizar_base_carga)

            # excel_etiquetas = gerar_etiquetas(tipo_filtro,tab_completa)

            # filenames.append(excel_etiquetas)

        if setor == 'Montagem':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].astype(str)

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')


            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('AM', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('AN', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('VJ', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('LC', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('VM', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('AV', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('CO', '')

            ###### retirando espaco em branco####

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa3', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa'] == '')].index, inplace=True)
            
            base_carretas = base_carretas.reset_index(drop=True)
            
            for i in range(len(base_carretas)):
                if len(base_carretas['Recurso'][i]) == 5:
                    base_carretas['Recurso'][i] = "0" + base_carretas['Recurso'][i]

            #### criando código único#####

            codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[6:10]

            #### filtrando data da carga#####

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == tipo_filtro)
            filtro_data = base_carga.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            filtro_data = filtro_data.reset_index(drop=True)
            filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

            for i in range(len(filtro_data)):
                if filtro_data['Recurso'][i][0] == '0':
                    filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]
                if len(filtro_data['Recurso'][i]) == 5:
                    filtro_data['Recurso'][i] = "0" + filtro_data['Recurso'][i]
            
            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            # base_carretas[base_carretas['Recurso'] == '034538M21']

            # carretas_agrupadas = filtro_data[['Recurso','Qtde']]
            # carretas_agrupadas = pd.DataFrame(filtro_data.groupby('Recurso').sum())
            # carretas_agrupadas = carretas_agrupadas[['Qtde']]

            # st.dataframe(carretas_agrupadas)

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa1 = tab_completa[['Código','Peca','Célula','Datas','Carga','Qtde_total']]

            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas','Carga']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            for d in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][d]) == 5:
                    tab_completa['Código'][d] = '0' + tab_completa['Código'][d]

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + tipo_filtro

            if idx == 0:
                st.write("Arquivos para download")

            k = 9

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            # print(tab_completa.columns)
            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas', 'Carga', 'PED_CHCRIACAO', 'Ano', 'codigo']).sum()
        
            tab_completa = tab_completa.reset_index(drop=True)

            # carga_unique = tab_completa['Carga'].unique()

            # for carga in carga_unique:
            file_counter = 1  # Contador de arquivos
            rows_per_file = 21  # Número máximo de linhas por arquivo
            k = 9  # Posição inicial no Excel

            for i in range(0, len(celulas_unique)):
                # Filtrar os dados para a célula atual
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel].reset_index(drop=True)

                # Verificar se o DataFrame está vazio
                if filtrar.empty:
                    continue

                # Índice inicial para controle de divisão em blocos
                start_index = 0

                while start_index < len(filtrar):
                    # Criar um novo Workbook para cada conjunto de 21 linhas
                    wb = Workbook()
                    wb = load_workbook('modelo_op_montagem.xlsx')
                    ws = wb.active

                    # Define o limite superior para as linhas deste arquivo
                    end_index = min(start_index + rows_per_file, len(filtrar))
                    k = 9  # Reseta a linha inicial para cada novo arquivo

                    # Escreve os dados no Excel para as linhas entre start_index e end_index
                    for j in range(start_index, end_index):
                        ws['G5'] = celulas_unique[0][i]  # Nome da célula
                        ws['AD5'] = hoje  # Data de hoje

                        # Gerar código único baseado na célula
                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C"
                        elif celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                        else:
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        # Preenchimento das células do Excel
                        ws['M4'] = tipo_filtro  # Data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k += 1

                    # Salvar o arquivo com numeração sequencial
                    file_name = f"Montagem {celulas_unique[0][i]} {data_nome_planilha} {file_counter}.xlsx"
                    wb.template = False
                    wb.save(file_name)
                    filenames.append(file_name)

                    # Incrementar o índice de início e o contador de arquivos
                    start_index = end_index
                    file_counter += 1

            # Preparar os dados para inserção no banco de dados
            data_formatada = datetime.strptime(tipo_filtro, '%d/%m/%Y').strftime('%Y-%m-%d')
            tab_completa['Datas'] = data_formatada
            data_insert_sql = tab_completa[['Célula', 'Código', 'Peca', 'Qtde_total', 'Datas']].values.tolist()

            # Chamar a função de inserção
            insert_montagem(data_formatada, data_insert_sql, check_atualizar_base_carga)

        if setor == 'Solda':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].astype(str)

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('CO', '')

            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('AM', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('AN', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('VJ', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('LC', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('VM', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('AV', '')
            base_carretas['Recurso'] = base_carretas['Recurso'].str.replace('CO', '')

            ###### retirando espaco em branco####

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa3', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa'] == '')].index, inplace=True)
            
            base_carretas = base_carretas.reset_index(drop=True)
            
            for i in range(len(base_carretas)):
                if len(base_carretas['Recurso'][i]) == 5:
                    base_carretas['Recurso'][i] = "0" + base_carretas['Recurso'][i]

            #### criando código único#####

            codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[6:10]

            #### filtrando data da carga#####

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == tipo_filtro)
            filtro_data = base_carga.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            filtro_data = filtro_data.reset_index(drop=True)
            filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

            for i in range(len(filtro_data)):
                if filtro_data['Recurso'][i][0] == '0':
                    filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]
                if len(filtro_data['Recurso'][i]) == 5:
                    filtro_data['Recurso'][i] = "0" + filtro_data['Recurso'][i]
            
            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            # base_carretas[base_carretas['Recurso'] == '034538M21']

            # carretas_agrupadas = filtro_data[['Recurso','Qtde']]
            # carretas_agrupadas = pd.DataFrame(filtro_data.groupby('Recurso').sum())
            # carretas_agrupadas = carretas_agrupadas[['Qtde']]

            # st.dataframe(carretas_agrupadas)

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa1 = tab_completa[['Código','Peca','Célula','Datas','Carga','Qtde_total']]

            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas','Carga']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            for d in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][d]) == 5:
                    tab_completa['Código'][d] = '0' + tab_completa['Código'][d]

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + tipo_filtro

            if idx == 0:
                st.write("Arquivos para download")

            k = 9

            # if carga_escolhida != 'Selecione':
            #     tab_completa = tab_completa[tab_completa['Carga'] == carga_escolhida]
            
            # print(tab_completa.columns)
            # tab_completa = tab_completa.groupby(
            #     ['Código', 'Peca', 'Célula', 'Datas', 'Carga', 'PED_CHCRIACAO', 'Ano', 'codigo']).sum()
        
            tab_completa = tab_completa.reset_index(drop=True)

            # carga_unique = tab_completa['Carga'].unique()

            # for carga in carga_unique:
            file_counter = 1  # Contador de arquivos
            rows_per_file = 21  # Número máximo de linhas por arquivo
            k = 9  # Posição inicial no Excel

            for i in range(0, len(celulas_unique)):
                # Filtrar os dados para a célula atual
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel].reset_index(drop=True)

                # Verificar se o DataFrame está vazio
                if filtrar.empty:
                    continue

                # Índice inicial para controle de divisão em blocos
                start_index = 0

                while start_index < len(filtrar):
                    # Criar um novo Workbook para cada conjunto de 21 linhas
                    wb = Workbook()
                    wb = load_workbook('modelo_op_solda.xlsx')
                    ws = wb.active

                    # Define o limite superior para as linhas deste arquivo
                    end_index = min(start_index + rows_per_file, len(filtrar))
                    k = 9  # Reseta a linha inicial para cada novo arquivo

                    # Escreve os dados no Excel para as linhas entre start_index e end_index
                    for j in range(start_index, end_index):
                        ws['G5'] = celulas_unique[0][i]  # Nome da célula
                        ws['AD5'] = hoje  # Data de hoje

                        # Gerar código único baseado na célula
                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C"
                        elif celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                        else:
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        # Preenchimento das células do Excel
                        ws['M4'] = tipo_filtro  # Data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k += 1

                    # Salvar o arquivo com numeração sequencial
                    file_name = f"Solda {celulas_unique[0][i]} {data_nome_planilha} {file_counter}.xlsx"
                    wb.template = False
                    wb.save(file_name)
                    filenames.append(file_name)

                    # Incrementar o índice de início e o contador de arquivos
                    start_index = end_index
                    file_counter += 1

            # Preparar os dados para inserção no banco de dados
            data_formatada = datetime.strptime(tipo_filtro, '%d/%m/%Y').strftime('%Y-%m-%d')
            tab_completa['Datas'] = data_formatada
            data_insert_sql = tab_completa[['Célula', 'Código', 'Peca', 'Qtde_total', 'Datas']].values.tolist()

        if setor == 'Serralheria':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')

            ###### retirando espaco em branco####

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa3', 'Etapa',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa4'] == '')].index, inplace=True)

            #### criando código único#####

            codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[6:10]

            #### filtrando data da carga#####

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == tipo_filtro)
            filtro_data = base_carga.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            for d in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][d]) == 5:
                    tab_completa['Código'][d] = '0' + tab_completa['Código'][d]

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + tipo_filtro

            if idx == 0:
                st.write("Arquivos para download")

            k = 9

            for i in range(0, len(celulas_unique)):

                wb = Workbook()
                wb = load_workbook('modelo_op_serralheria.xlsx')
                ws = wb.active

                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel]
                filtrar.reset_index(inplace=True)
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])

                if len(filtrar) > 21:

                    for j in range(0, 21):

                        ws['G5'] = celulas_unique[0][i]  # nome da coluna é '0'
                        ws['AD5'] = hoje  # data de hoje
                        # código único para cada sequenciamento
                        ws['AK4'] = celulas_unique[0][i][0:3] + \
                            codigo_unico + celulas_unique[0][i][:4]

                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "C"

                        if celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "S"

                        else:
                            # código único para cada sequenciamento
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        filtrar = tab_completa.loc[filtro_excel]
                        filtrar.reset_index(inplace=True)

                        ws['M4'] = tipo_filtro  # data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k = k + 1

                        wb.template = False
                        wb.save('Serralheria ' + celulas_unique[0][i] + data_nome_planilha +'1.xlsx')

                    my_file = "Serralheria " + celulas_unique[0][i] + data_nome_planilha +'1.xlsx'
                    filenames.append(my_file)

                    k = 9

                    wb = Workbook()
                    wb = load_workbook('modelo_op_serralheria.xlsx')
                    ws = wb.active

                    filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                    filtrar = tab_completa.loc[filtro_excel]
                    filtrar.reset_index(inplace=True)
                    filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])

                    if len(filtrar) > 21:

                        j = 21

                        for j in range(21, len(filtrar)):

                            ws['G5'] = celulas_unique[0][i]  # nome da coluna é '0'
                            ws['AD5'] = hoje  # data de hoje

                            if celulas_unique[0][i] == "EIXO COMPLETO":
                                ws['AK4'] = celulas_unique[0][i][0:3] + \
                                    codigo_unico + "C"

                            if celulas_unique[0][i] == "EIXO SIMPLES":
                                ws['AK4'] = celulas_unique[0][i][0:3] + \
                                    codigo_unico + "S"

                            else:
                                # código único para cada sequenciamento
                                ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                            filtrar = tab_completa.loc[filtro_excel]
                            filtrar.reset_index(inplace=True)

                            ws['M4'] = tipo_filtro  # data da carga
                            ws['B' + str(k)] = filtrar['Código'][j]
                            ws['G' + str(k)] = filtrar['Peca'][j]
                            ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                            k = k + 1

                            wb.template = False
                            wb.save('Serralheria ' +
                                    celulas_unique[0][i] + '.xlsx')

                        my_file = "Serralheria " + celulas_unique[0][i] + data_nome_planilha +'.xlsx'
                        filenames.append(my_file)

                else:

                    j = 0
                    k = 9

                    for j in range(0, 21-(21-len(filtrar))):

                        ws['G5'] = celulas_unique[0][i]  # nome da coluna é '0'
                        ws['AD5'] = hoje  # data de hoje

                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "C"

                        if celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "S"

                        else:

                            # código único para cada sequenciamento
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        filtrar = tab_completa.loc[filtro_excel]
                        filtrar.reset_index(inplace=True)

                        ws['M4'] = tipo_filtro  # data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k = k + 1

                        wb.template = False
                        wb.save('Serralheria ' + celulas_unique[0][i] + data_nome_planilha +'.xlsx')

                    k = 9

                    my_file = "Serralheria " + celulas_unique[0][i] + data_nome_planilha +'.xlsx'
                    filenames.append(my_file)

            name_sheet4 = 'Base gerador de ordem de producao'
            worksheet4 = 'Serralheria'

            sh = sa.open(name_sheet4)
            wks4 = sh.worksheet(worksheet4)

            list4 = wks4.get_all_records()
            table = pd.DataFrame(list4)

            table = table.astype(str)

            for i in range(len(table)):
                if len(table['CODIGO'][i]) == 5:
                    table['CODIGO'][i] = '0'+table['CODIGO'][i]

            tab_completa['Carimbo'] = tipo_filtro + 'Serralheria'
            tab_completa['Data_carga'] = tipo_filtro

            tab_completa1 = tab_completa[[
                'Carimbo', 'Célula', 'Código', 'Peca', 'Qtde_total', 'Data_carga']]
            tab_completa1['Data_carga'] = tipo_filtro
            tab_completa1['Setor'] = 'Serralheria'

            tab_completa_2 = tab_completa1

            table = table.loc[(table.UNICO == tipo_filtro + 'Serralheria')]

            list_columns = table.columns.values.tolist()

            tab_completa_2.columns = list_columns

            frames = [table, tab_completa_2]

            table = pd.concat(frames)
            table['QT_ITENS'] = table['QT_ITENS'].astype(int)
            table = table.drop_duplicates(keep=False)

            tab_completa1 = table.values.tolist()
            sh.values_append('Serralheria', {'valueInputOption': 'RAW'}, {
                            'values': tab_completa1})

        if setor == 'Carpintaria':

            base_carretas['Código'] = base_carretas['Código'].astype(str)
            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AN', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VJ', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('LC', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('VM', '')
            base_carga['Recurso'] = base_carga['Recurso'].str.replace('AV', '')

            ###### retirando espaco em branco####

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carretas.drop(['Etapa2', 'Etapa3', 'Etapa',
                            'Etapa4'], axis=1, inplace=True)

            # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
            base_carretas.drop(
                base_carretas[(base_carretas['Etapa5'] == '')].index, inplace=True)

            #### criando código único#####

            codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[6:10]

            #### filtrando data da carga#####

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == tipo_filtro)
            filtro_data = base_carga.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carretas[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            for d in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][d]) == 5:
                    tab_completa['Código'][d] = '0' + tab_completa['Código'][d]

            # criando coluna de código para arquivar

            hoje = datetime.now()

            ts = pd.Timestamp(hoje)

            hoje1 = hoje.strftime('%d%m%Y')

            controle_seq = tab_completa
            controle_seq["codigo"] = hoje1 + tipo_filtro

            if idx == 0:
                st.write("Arquivos para download")

            k = 9

            for i in range(0, len(celulas_unique)):

                wb = Workbook()
                wb = load_workbook('modelo_op_carpintaria.xlsx')
                ws = wb.active

                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel]
                filtrar.reset_index(inplace=True)
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])

                if len(filtrar) > 21:

                    for j in range(0, 21):

                        ws['G5'] = celulas_unique[0][i]  # nome da coluna é '0'
                        ws['AD5'] = hoje  # data de hoje
                        # código único para cada sequenciamento
                        ws['AK4'] = celulas_unique[0][i][0:3] + \
                            codigo_unico + celulas_unique[0][i][:4]

                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "C"

                        if celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "S"

                        else:
                            # código único para cada sequenciamento
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        filtrar = tab_completa.loc[filtro_excel]
                        filtrar.reset_index(inplace=True)

                        ws['M4'] = tipo_filtro  # data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k = k + 1

                        wb.template = False
                        wb.save('Carpintaria ' + celulas_unique[0][i] + data_nome_planilha +'1.xlsx')

                    my_file = "Carpintaria " + celulas_unique[0][i] + data_nome_planilha +'1.xlsx'
                    filenames.append(my_file)

                    k = 9

                    wb = Workbook()
                    wb = load_workbook('modelo_op_carpintaria.xlsx')
                    ws = wb.active

                    filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                    filtrar = tab_completa.loc[filtro_excel]
                    filtrar.reset_index(inplace=True)
                    filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])

                    if len(filtrar) > 21:

                        j = 21

                        for j in range(21, len(filtrar)):

                            ws['G5'] = celulas_unique[0][i]  # nome da coluna é '0'
                            ws['AD5'] = hoje  # data de hoje

                            if celulas_unique[0][i] == "EIXO COMPLETO":
                                ws['AK4'] = celulas_unique[0][i][0:3] + \
                                    codigo_unico + "C"

                            if celulas_unique[0][i] == "EIXO SIMPLES":
                                ws['AK4'] = celulas_unique[0][i][0:3] + \
                                    codigo_unico + "S"

                            else:
                                # código único para cada sequenciamento
                                ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                            filtrar = tab_completa.loc[filtro_excel]
                            filtrar.reset_index(inplace=True)

                            ws['M4'] = tipo_filtro  # data da carga
                            ws['B' + str(k)] = filtrar['Código'][j]
                            ws['G' + str(k)] = filtrar['Peca'][j]
                            ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                            k = k + 1

                            wb.template = False
                            wb.save('Carpintaria ' +
                                    celulas_unique[0][i] + data_nome_planilha +'.xlsx')

                        my_file = "Carpintaria " + celulas_unique[0][i] + data_nome_planilha +'.xlsx'
                        filenames.append(my_file)

                else:

                    j = 0
                    k = 9

                    for j in range(0, 21-(21-len(filtrar))):

                        ws['G5'] = celulas_unique[0][i]  # nome da coluna é '0'
                        ws['AD5'] = hoje  # data de hoje

                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "C"

                        if celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + \
                                codigo_unico + "S"

                        else:

                            # código único para cada sequenciamento
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico

                        filtrar = tab_completa.loc[filtro_excel]
                        filtrar.reset_index(inplace=True)

                        ws['M4'] = tipo_filtro  # data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k = k + 1

                        wb.template = False
                        wb.save('Carpintaria ' + celulas_unique[0][i] + data_nome_planilha +'.xlsx')

                    k = 9

                    my_file = "Carpintaria " + celulas_unique[0][i] + data_nome_planilha +'.xlsx'
                    filenames.append(my_file)

            name_sheet4 = 'Base gerador de ordem de producao'
            worksheet4 = 'Carpintaria'

            sh = sa.open(name_sheet4)
            wks4 = sh.worksheet(worksheet4)

            list4 = wks4.get_all_records()
            table = pd.DataFrame(list4)

            table = table.astype(str)

            for i in range(len(table)):
                if len(table['CODIGO'][i]) == 5:
                    table['CODIGO'][i] = '0'+table['CODIGO'][i]

            tab_completa['Carimbo'] = tipo_filtro + 'Carpintaria'
            tab_completa['Data_carga'] = tipo_filtro

            tab_completa1 = tab_completa[[
                'Carimbo', 'Célula', 'Código', 'Peca', 'Qtde_total', 'Data_carga']]
            tab_completa1['Data_carga'] = tipo_filtro
            tab_completa1['Setor'] = 'Carpintaria'

            tab_completa_2 = tab_completa1

            table = table.loc[(table.UNICO == tipo_filtro + 'Carpintaria')]

            list_columns = table.columns.values.tolist()

            tab_completa_2.columns = list_columns

            frames = [table, tab_completa_2]

            table = pd.concat(frames)
            table['QT_ITENS'] = table['QT_ITENS'].astype(int)
            table = table.drop_duplicates(keep=False)

            tab_completa1 = table.values.tolist()
            sh.values_append('Carpintaria', {'valueInputOption': 'RAW'}, {
                            'values': tab_completa1})

        if setor == 'Etiquetas':
            
            base_carreta_montagem = base_carretas.copy()
            base_carga_montagem = base_carga.copy()

            ########################### Montagem ###########################

            base_carreta_montagem['Código'] = base_carreta_montagem['Código'].astype(str)
            base_carreta_montagem['Recurso'] = base_carreta_montagem['Recurso'].astype(str)

            ####### retirando cores dos códigos######

            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].astype(str)

            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('AM', '')
            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('AN', '')
            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('VJ', '')
            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('LC', '')
            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('VM', '')
            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.replace('AV', '')

            # base_carga_montagem[base_carga_montagem['Datas'] == '01/06/2023']

            ###### retirando espaco em branco####

            base_carga_montagem['Recurso'] = base_carga_montagem['Recurso'].str.strip()

            ##### excluindo colunas e linhas#####

            base_carreta_montagem.drop(['Etapa2', 'Etapa3', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            # & (base_carreta_montagem['Unit_Price'] < 600)].index, inplace=True)
            base_carreta_montagem.drop(
                base_carreta_montagem[(base_carreta_montagem['Etapa'] == '')].index, inplace=True)

            #### criando código único#####

            codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[6:10]

            #### filtrando data da carga#####

            datas_unique = pd.DataFrame(base_carga_montagem['Datas'].unique())

            escolha_data = (base_carga_montagem['Datas'] == tipo_filtro)
            filtro_data = base_carga_montagem.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            filtro_data = filtro_data.reset_index(drop=True)
            filtro_data['Recurso'] = filtro_data['Recurso'].astype(str)

            for i in range(len(filtro_data)):
                if filtro_data['Recurso'][i][0] == '0':
                    filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]

            ##### juntando planilhas de acordo com o recurso#######

            tab_completa = pd.merge(filtro_data, base_carreta_montagem[[
                                    'Recurso', 'Código', 'Peca', 'Qtde', 'Célula']], on=['Recurso'], how='left')
            tab_completa = tab_completa.dropna(axis=0)

            # carretas_agrupadas = filtro_data[['Recurso','Qtde']]
            # carretas_agrupadas = pd.DataFrame(filtro_data.groupby('Recurso').sum())
            # carretas_agrupadas = carretas_agrupadas[['Qtde']]

            # st.dataframe(carretas_agrupadas)

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa.reset_index(inplace=True, drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(inplace=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # criando coluna de quantidade total de itens

            try:
                tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(
                    ',', '.')
            except:
                pass

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas']).sum()

            # tab_completa = tab_completa.drop_duplicates()

            tab_completa.reset_index(inplace=True)

            # tratando coluna de código e recurso

            for d in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][d]) == 5:
                    tab_completa['Código'][d] = '0' + tab_completa['Código'][d]
            
            tab_completa = tab_completa[tab_completa['Célula'] != 'CHASSI']

            # tab_completa['Código'] = 'Montagem/Solda ' + tab_completa['Código']
            tab_completa_montagem = tab_completa.copy()
            # teste1 = tab_completa_montagem[tab_completa_montagem['Célula'] == 'CHASSI']
            # teste2 = tab_completa[tab_completa['Célula'] == 'CHASSI']
            
            # pd.concat([teste2,teste1])

            chassi_separado = tratar_conjuntos_iguais(base_carretas,base_carga)

            tab_completa_montagem = chassi_separado.copy()

            # tab_completa_montagem = pd.concat([tab_completa_montagem,chassi_separado])

            ########################### Pintura ###########################

            base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)

            base_carretas.drop(['Etapa', 'Etapa3', 'Etapa4',
                            'Etapa5'], axis=1, inplace=True)

            base_carretas.drop(
                base_carretas[(base_carretas['Etapa2'] == '')].index, inplace=True)

            base_carretas = base_carretas.reset_index(drop=True)

            base_carretas = base_carretas.astype(str)

            for d in range(0, base_carretas.shape[0]):

                if len(base_carretas['Código'][d]) == 5:
                    base_carretas['Código'][d] = '0' + base_carretas['Código'][d]

            # separando string por "-" e adicionando no dataframe antigo

            base_carga["Recurso"] = base_carga["Recurso"].astype(str)

            tratando_coluna = base_carga["Recurso"].str.split(
                " - ", n=1, expand=True)

            base_carga['Recurso'] = tratando_coluna[0]

            # tratando cores da string

            base_carga['Recurso_cor'] = base_carga['Recurso']

            base_carga = base_carga.reset_index(drop=True)

            df_cores = pd.DataFrame({'Recurso_cor': ['AN', 'LJ', 'VJ', 'LC', 'VM', 'AV', 'AS', 'sem_cor'],
                                    'cor': ['Azul', 'Verde', 'Laranja Jacto','Laranja', 'Vermelho', 'Amarelo', 'Laranja']})

            cores = ['AM', 'AN', 'VJ', 'LC', 'VM', 'AV', 'AS']

            base_carga = base_carga.astype(str)

            for r in range(0, base_carga.shape[0]):
                base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][len(
                    base_carga['Recurso_cor'][r])-3:len(base_carga['Recurso_cor'][r])]
                base_carga['Recurso_cor'] = base_carga['Recurso_cor'].str.strip()

                if len(base_carga['Recurso_cor'][r]) > 2:
                    base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][1:3]

                if base_carga['Recurso_cor'][r] not in cores:
                    base_carga['Recurso_cor'][r] = "LC"

            base_carga = pd.merge(base_carga, df_cores, on=[
                                'Recurso_cor'], how='left')

            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'AN', '')  # Azul
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'VJ', '')  # Verde
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'LC', '')  # Laranja
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'VM', '')  # Vermelho
            base_carga['Recurso'] = base_carga['Recurso'].str.replace(
                'AV', '')  # Amarelo

            base_carga['Recurso'] = base_carga['Recurso'].str.strip()

            datas_unique = pd.DataFrame(base_carga['Datas'].unique())

            escolha_data = (base_carga['Datas'] == tipo_filtro)
            filtro_data = base_carga.loc[escolha_data]
            filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)

            # procv e trazendo as colunas que quero ver

            filtro_data = filtro_data.reset_index(drop=True)

            for i in range(len(filtro_data)):
                if filtro_data['Recurso'][i][0] == '0':
                    filtro_data['Recurso'][i] = filtro_data['Recurso'][i][1:]

            tab_completa = pd.merge(filtro_data, base_carretas, on=[
                                    'Recurso'], how='left')

            tab_completa['Código'] = tab_completa['Código'].astype(str)

            tab_completa = tab_completa.reset_index(drop=True)

            celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
            celulas_unique = celulas_unique.dropna(axis=0)
            celulas_unique.reset_index(drop=True)

            recurso_unique = pd.DataFrame(tab_completa['Recurso'].unique())
            recurso_unique = recurso_unique.dropna(axis=0)

            # tratando coluna de código

            for t in range(0, tab_completa.shape[0]):

                if len(tab_completa['Código'][t]) == 5:
                    tab_completa['Código'][t] = '0' + \
                        tab_completa['Código'][t][0:5]

                if len(tab_completa['Código'][t]) == 8:
                    tab_completa['Código'][t] = tab_completa['Código'][t][0:6]

            # criando coluna de quantidade total de itens

            tab_completa = tab_completa.dropna()

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(',', '.')

            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

            tab_completa = tab_completa.dropna(axis=0)

            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
            tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

            tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * \
                tab_completa['Qtde_y']

            tab_completa = tab_completa.drop(
                columns=['Recurso', 'Qtde_x', 'Qtde_y', 'LEAD TIME', 'flag peça', 'Etapa2'])

            tab_completa = tab_completa.groupby(
                ['Código', 'Peca', 'Célula', 'Datas', 'Recurso_cor', 'cor']).sum()
            tab_completa.reset_index(inplace=True)

            # linha abaixo exclui eixo simples do sequenciamento da pintura
            # tab_completa.drop(tab_completa.loc[tab_completa['Célula']=='EIXO SIMPLES'].index, inplace=True)
            tab_completa.reset_index(inplace=True, drop=True)

            for t in range(0, len(tab_completa)):

                if tab_completa['Célula'][t] == 'FUEIRO' or \
                        tab_completa['Célula'][t] == 'LATERAL' or \
                        tab_completa['Célula'][t] == 'PLAT. TANQUE. CAÇAM.':

                    tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + \
                        tab_completa['Recurso_cor'][t]

                else:

                    tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + 'CO'
                    tab_completa['cor'][t] = 'Cinza'

            # Consumo de tinta

            # tab_completa = tab_completa.merge(df_consumo_pu[['Codigo item','Consumo Pó (kg)','Consumo PU (L)','Consumo Catalisador (L)']], left_on='Código', right_on='Codigo item', how='left').fillna(0)
            
            # tab_completa['Consumo Pó (kg)'] = tab_completa['Consumo Pó (kg)'] * tab_completa['Qtde_total']
            # tab_completa['Consumo PU (L)'] = tab_completa['Consumo PU (L)'] * tab_completa['Qtde_total']
            # tab_completa['Consumo Catalisador (L)'] = tab_completa['Consumo Catalisador (L)'] * tab_completa['Qtde_total']

            # consumo_po = sum(tab_completa['Consumo Pó (kg)'])
            # consumo_po = f'{round(consumo_po / 25, 2)} caixa(s)'

            # consumo_pu_litros = sum(tab_completa['Consumo Pó (kg)'])
            # consumo_pu_latas = round(consumo_pu_litros / 3.08, 2)
            # consumo_pu = f'{consumo_pu_latas} lata(s)'

            # consumo_catalisador_litros = sum(tab_completa['Consumo Catalisador (L)'])
            # consumo_catalisador_latas = round(consumo_catalisador_litros * 1000 / 400, 2)
            # consumo_cata = f'{consumo_catalisador_latas} lata(s)'

            # diluente = f'{round((consumo_pu_litros * 0.80) / 5, 2)} lata(s)'

            tab_completa = tab_completa.groupby(['Código','Peca','Célula','Datas','cor','Recurso_cor']).sum('Qtde_total').reset_index()
            tab_completa = tab_completa[tab_completa['Célula'] != 'CHASSI']
            # tab_completa = pd.concat([tab_completa,tab_completa_montagem[tab_completa_montagem['Célula'] == 'CHASSI']])
            # tab_completa['cor'] = tab_completa['cor'].fillna('Montagem/Solda')
            # tab_completa['Recurso_cor'] = tab_completa['Recurso_cor'].fillna(tab_completa['Código'] + ' ' + tab_completa['cor'])
            ###########################################################################################

            # cor_unique = tab_completa['cor'].unique()

            # st.write("Arquivos para download")

            gerar_etiquetas(tipo_filtro,tab_completa,tab_completa_montagem)
            
            st.write("Etiquetas adicionada na planilha: https://docs.google.com/spreadsheets/d/1jojKHPBKeALheutJyphsPS-LGNu1e2BC54AAqRnF-us/edit#gid=1389272651")

            # filenames.append(excel_etiquetas)
        
        # data_resumo = tipo_filtro.strftime("%d/%m/%Y")
        st.write(f"Resumo: {tipo_filtro}")
        base_carga_filtro = base_carga.query("Datas == @tipo_filtro")
        base_carga_filtro.dropna(inplace=True)
        base_carga_filtro = base_carga_filtro[base_carga_filtro['Qtde'] != '']
        base_carga_filtro = base_carga_filtro[['Recurso', 'Qtde']]
        base_carga_filtro['Qtde'] = base_carga_filtro['Qtde'].astype(int)
        base_carga_filtro = base_carga_filtro.groupby('Recurso').sum()

        try:
            tab_completa[['Datas','Célula', 'Código', 'Peca', 'Qtde_total', 'cor']]
        except:
            tab_completa[['Célula','Código','Peca','Qtde_total']]    

    if len(filenames)!=0:

        filenames_unique = list(set(filenames))

        with zipfile.ZipFile("Arquivos.zip", mode="w") as archive:
            for filename in filenames_unique:
                archive.write(filename)

        with open("Arquivos.zip", "rb") as fp:
            btn = st.download_button(
                label="Download arquivos",
                data=fp,
                file_name="Arquivos.zip",
                mime="application/zip"
            )
    
    else:
        pass